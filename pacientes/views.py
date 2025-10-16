from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.core.paginator import Paginator
from django.db.models import Q
from django.db import transaction
from django.http import HttpResponse
from django.template.loader import get_template
from xhtml2pdf import pisa
from io import BytesIO
import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

from .models import Paciente, Pais, Estado, Ciudad, Direccion, Telefono
from .forms import PacienteForm, DireccionFormSet, TelefonoFormSet
from sistema_medico.settings import BASE_DIR

# --- Vistas CRUD y de Búsqueda --- #

def index(request):
    pacientes_list = Paciente.objects.all().order_by('apellido', 'nombre')
    paginator = Paginator(pacientes_list, 10)
    page_number = request.GET.get('page')
    pacientes = paginator.get_page(page_number)
    return render(request, 'pacientes/index.html', {'pacientes': pacientes})

@transaction.atomic
def create(request):
    if request.method == 'POST':
        paciente_form = PacienteForm(request.POST)
        direccion_formset = DireccionFormSet(request.POST)
        telefono_formset = TelefonoFormSet(request.POST)
        if paciente_form.is_valid() and direccion_formset.is_valid() and telefono_formset.is_valid():
            paciente = paciente_form.save()
            direccion_formset.instance = paciente
            direccion_formset.save()
            telefono_formset.instance = paciente
            telefono_formset.save()
            messages.success(request, 'Paciente creado correctamente.')
            return redirect('pacientes:index')
        else:
            messages.error(request, 'Por favor, corrija los errores en el formulario.')
    else:
        paciente_form = PacienteForm()
        direccion_formset = DireccionFormSet()
        telefono_formset = TelefonoFormSet()
    paises = Pais.objects.all().order_by('nombre')
    return render(request, 'pacientes/create.html', {
        'form': paciente_form,
        'direccion_formset': direccion_formset,
        'telefono_formset': telefono_formset,
        'paises': paises,
    })

def show(request, paciente_id):
    paciente = get_object_or_404(Paciente, id=paciente_id)
    direcciones = Direccion.objects.filter(paciente=paciente).select_related('ciudad__estado__pais')
    telefonos = Telefono.objects.filter(paciente=paciente).select_related('tipo_telefono')
    return render(request, 'pacientes/show.html', {
        'paciente': paciente,
        'direcciones': direcciones,
        'telefonos': telefonos
    })

@transaction.atomic
def edit(request, paciente_id):
    paciente = get_object_or_404(Paciente, id=paciente_id)
    if request.method == 'POST':
        paciente_form = PacienteForm(request.POST, instance=paciente)
        direccion_formset = DireccionFormSet(request.POST, instance=paciente)
        telefono_formset = TelefonoFormSet(request.POST, instance=paciente)
        if paciente_form.is_valid() and direccion_formset.is_valid() and telefono_formset.is_valid():
            paciente_form.save()
            direccion_formset.save()
            telefono_formset.save()
            messages.success(request, 'Paciente actualizado correctamente.')
            return redirect('pacientes:index')
    else:
        paciente_form = PacienteForm(instance=paciente)
        direccion_formset = DireccionFormSet(instance=paciente)
        telefono_formset = TelefonoFormSet(instance=paciente)
    paises = Pais.objects.all().order_by('nombre')
    return render(request, 'pacientes/edit.html', {
        'form': paciente_form,
        'direccion_formset': direccion_formset,
        'telefono_formset': telefono_formset,
        'paciente': paciente,
        'paises': paises,
    })

@transaction.atomic
def destroy(request, paciente_id):
    paciente = get_object_or_404(Paciente, id=paciente_id)
    if request.method == 'POST':
        paciente_info = f"{paciente.nombre} {paciente.apellido}"
        paciente.delete()
        messages.success(request, f'Paciente {paciente_info} eliminado correctamente.')
        return redirect('pacientes:index')
    return render(request, 'pacientes/destroy.html', {'paciente': paciente})

def search(request):
    query = request.GET.get('q', '')
    pacientes = Paciente.objects.all().order_by('apellido', 'nombre')
    if query:
        pacientes = pacientes.filter(Q(nombre__icontains=query) | Q(apellido__icontains=query) | Q(numero_documento__icontains=query))
    paginator = Paginator(pacientes, 10)
    page_number = request.GET.get('page')
    pacientes_page = paginator.get_page(page_number)
    return render(request, 'pacientes/search.html', {'pacientes': pacientes_page, 'query': query})

# --- Vistas AJAX --- #

def cargar_estados(request):
    pais_id = request.GET.get('pais_id')
    estados = Estado.objects.filter(pais_id=pais_id).order_by('nombre')
    return render(request, 'pacientes/dropdown_list_options.html', {'opciones': estados})

def cargar_ciudades(request):
    estado_id = request.GET.get('estado_id')
    ciudades = Ciudad.objects.filter(estado_id=estado_id).order_by('nombre')
    return render(request, 'pacientes/dropdown_list_options.html', {'opciones': ciudades})

# --- Vistas de Exportación --- #

def exportar_pacientes_pdf(request):
    pacientes = Paciente.objects.all().select_related('direccion__ciudad__estado__pais').prefetch_related('telefonos__tipo_telefono').order_by('apellido', 'nombre')
    logo_path = str(BASE_DIR / 'static/img/logo.png')
    
    context = {
        'pacientes': pacientes,
        'logo_path': logo_path,
        'generation_date': datetime.datetime.now().strftime('%d/%m/%Y %H:%M')
    }

    template = get_template('pacientes/pdf_template.html')
    html = template.render(context)
    result = BytesIO()
    pdf = pisa.pisaDocument(BytesIO(html.encode("UTF-8")), result)

    if not pdf.err:
        response = HttpResponse(result.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="dossier_pacientes_{}.pdf"'.format(datetime.datetime.now().strftime("%Y%m%d"))
        return response
    
    return HttpResponse("Error al generar el PDF.", status=400)

def exportar_pacientes_excel(request):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="listado_pacientes_{}.xlsx"'.format(datetime.datetime.now().strftime("%Y%m%d"))
    wb = Workbook()
    ws = wb.active
    ws.title = "Pacientes"
    ws.freeze_panes = 'A2'
    headers = ['Cédula', 'Nombre', 'Apellido', 'Fecha de Nacimiento', 'Edad', 'Género', 'Email', 'Dirección', 'Ciudad', 'Estado', 'País', 'Teléfonos']
    ws.append(headers)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="0d6efd", end_color="0d6efd", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    pacientes = Paciente.objects.all().select_related('direccion__ciudad__estado__pais').prefetch_related('telefonos').order_by('apellido', 'nombre')
    for paciente in pacientes:
        direccion_obj = paciente.direccion
        dir_completa, ciudad, estado, pais = "N/A", "N/A", "N/A", "N/A"
        if direccion_obj:
            dir_completa = direccion_obj.direccion
            if direccion_obj.ciudad:
                ciudad = direccion_obj.ciudad.nombre
                if direccion_obj.ciudad.estado:
                    estado = direccion_obj.ciudad.estado.nombre
                    if direccion_obj.ciudad.estado.pais:
                        pais = direccion_obj.ciudad.estado.pais.nombre
        telefonos = ", ".join([t.numero for t in paciente.telefonos.all()])
        ws.append([
            paciente.numero_documento, paciente.nombre, paciente.apellido, paciente.fecha_nacimiento, paciente.edad,
            paciente.get_genero_display(), paciente.email or 'N/A', dir_completa, ciudad, estado, pais, telefonos
        ])
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width
    wb.save(response)
    return response