from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.core.paginator import Paginator
from django.db.models import Q
from django.utils import timezone
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_POST
from django.template.loader import get_template
from xhtml2pdf import pisa
from io import BytesIO
import datetime
from sistema_medico.settings import BASE_DIR
import json

from django.contrib.auth.decorators import login_required

from core.decorators import personal_medico_required

from .models import Categoria, Proveedor, Medicamento, Inventario, MovimientoInventario
from .forms import CategoriaForm, ProveedorForm, MedicamentoForm, InventarioForm, MedicamentoModalForm, CategoriaModalForm, ProveedorModalForm, MovimientoSalidaForm

# Vistas para Categorías
@personal_medico_required
def listar_categorias(request):
    categorias_list = Categoria.objects.all().order_by('nombre')
    paginator = Paginator(categorias_list, 10)
    page_number = request.GET.get('page')
    categorias = paginator.get_page(page_number)
    return render(request, 'inventario/categorias/listar.html', {'categorias': categorias})

@personal_medico_required
def crear_categoria(request):
    if request.method == 'POST':
        form = CategoriaForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Categoría creada exitosamente.')
            return redirect('inventario:listar_categorias')
    else:
        form = CategoriaForm()
    return render(request, 'inventario/categorias/formulario.html', {'form': form, 'titulo': 'Crear Categoría'})

@personal_medico_required
def editar_categoria(request, categoria_id):
    categoria = get_object_or_404(Categoria, id=categoria_id)
    if request.method == 'POST':
        form = CategoriaForm(request.POST, instance=categoria)
        if form.is_valid():
            form.save()
            messages.success(request, 'Categoría actualizada exitosamente.')
            return redirect('inventario:listar_categorias')
    else:
        form = CategoriaForm(instance=categoria)
    return render(request, 'inventario/categorias/formulario.html', {'form': form, 'titulo': 'Editar Categoría'})

@personal_medico_required
def eliminar_categoria(request, categoria_id):
    categoria = get_object_or_404(Categoria, id=categoria_id)
    if request.method == 'POST':
        categoria.delete()
        messages.success(request, 'Categoría eliminada exitosamente.')
        return redirect('inventario:listar_categorias')
    return render(request, 'inventario/categorias/eliminar.html', {'categoria': categoria})

# Vistas para Proveedores
@personal_medico_required
def listar_proveedores(request):
    proveedores_list = Proveedor.objects.all().order_by('nombre')
    paginator = Paginator(proveedores_list, 10)
    page_number = request.GET.get('page')
    proveedores = paginator.get_page(page_number)
    return render(request, 'inventario/proveedores/listar.html', {'proveedores': proveedores})

@personal_medico_required
def crear_proveedor(request):
    if request.method == 'POST':
        form = ProveedorForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Proveedor creado exitosamente.')
            return redirect('inventario:listar_proveedores')
    else:
        form = ProveedorForm()
    return render(request, 'inventario/proveedores/formulario.html', {'form': form, 'titulo': 'Crear Proveedor'})

@personal_medico_required
def editar_proveedor(request, proveedor_id):
    proveedor = get_object_or_404(Proveedor, id=proveedor_id)
    if request.method == 'POST':
        form = ProveedorForm(request.POST, instance=proveedor)
        if form.is_valid():
            form.save()
            messages.success(request, 'Proveedor actualizado exitosamente.')
            return redirect('inventario:listar_proveedores')
    else:
        form = ProveedorForm(instance=proveedor)
    return render(request, 'inventario/proveedores/formulario.html', {'form': form, 'titulo': 'Editar Proveedor'})

@personal_medico_required
def eliminar_proveedor(request, proveedor_id):
    proveedor = get_object_or_404(Proveedor, id=proveedor_id)
    if request.method == 'POST':
        proveedor.delete()
        messages.success(request, 'Proveedor eliminado exitosamente.')
        return redirect('inventario:listar_proveedores')
    return render(request, 'inventario/proveedores/eliminar.html', {'proveedor': proveedor})

# Vistas para Medicamentos
@personal_medico_required
def listar_medicamentos(request):
    medicamentos_list = Medicamento.objects.select_related('categoria', 'proveedor').all().order_by('nombre')
    paginator = Paginator(medicamentos_list, 10)
    page_number = request.GET.get('page')
    medicamentos = paginator.get_page(page_number)
    return render(request, 'inventario/medicamentos/listar.html', {'medicamentos': medicamentos})

@personal_medico_required
def crear_medicamento(request):
    if request.method == 'POST':
        form = MedicamentoForm(request.POST)
        if form.is_valid():
            medicamento = form.save(commit=False)
            medicamento.save()
            messages.success(request, 'Medicamento creado exitosamente.')
            return redirect('inventario:listar_medicamentos')
    else:
        form = MedicamentoForm()
    return render(request, 'inventario/medicamentos/formulario.html', {'form': form, 'titulo': 'Crear Medicamento'})

@personal_medico_required
def editar_medicamento(request, medicamento_id):
    medicamento = get_object_or_404(Medicamento, id=medicamento_id)
    if request.method == 'POST':
        form = MedicamentoForm(request.POST, instance=medicamento)
        if form.is_valid():
            form.save()
            messages.success(request, 'Medicamento actualizado exitosamente.')
            return redirect('inventario:listar_medicamentos')
    else:
        form = MedicamentoForm(instance=medicamento)
    return render(request, 'inventario/medicamentos/formulario.html', {'form': form, 'titulo': 'Editar Medicamento'})

@personal_medico_required
def eliminar_medicamento(request, medicamento_id):
    medicamento = get_object_or_404(Medicamento, id=medicamento_id)
    if request.method == 'POST':
        medicamento.delete()
        messages.success(request, 'Medicamento eliminado exitosamente.')
        return redirect('inventario:listar_medicamentos')
    return render(request, 'inventario/medicamentos/eliminar.html', {'medicamento': medicamento})

# Vistas para Inventario
@personal_medico_required
def listar_inventario(request):
    inventario_list = Inventario.objects.select_related('medicamento').all().order_by('-created_at')
    paginator = Paginator(inventario_list, 10)
    page_number = request.GET.get('page')
    inventario = paginator.get_page(page_number)
    return render(request, 'inventario/inventario/listar.html', {
        'inventario': inventario,
        'today': timezone.now().date()
    })

@personal_medico_required
def crear_inventario(request):
    if request.method == 'POST':
        form = InventarioForm(request.POST)
        if form.is_valid():
            # Guardar la existencia de inventario
            inventario = form.save()
            
            # Crear el movimiento de inventario correspondiente
            MovimientoInventario.objects.create(
                medicamento=inventario.medicamento,
                tipo='entrada',
                cantidad=inventario.cantidad,
                descripcion=f"Ingreso de lote #{inventario.lote}",
                usuario=request.user.username
            )
            
            messages.success(request, 'Existencia de inventario creada y movimiento de entrada registrado exitosamente.')
            return redirect('inventario:listar_inventario')
    else:
        form = InventarioForm()
    return render(request, 'inventario/inventario/formulario.html', {'form': form, 'titulo': 'Agregar Existencia'})

@personal_medico_required
def editar_inventario(request, inventario_id):
    inventario = get_object_or_404(Inventario, id=inventario_id)
    if request.method == 'POST':
        form = InventarioForm(request.POST, instance=inventario)
        if form.is_valid():
            form.save()
            messages.success(request, 'Existencia de inventario actualizada exitosamente.')
            return redirect('inventario:listar_inventario')
    else:
        form = InventarioForm(instance=inventario)
    return render(request, 'inventario/inventario/formulario.html', {'form': form, 'titulo': 'Editar Existencia'})

@personal_medico_required
def eliminar_inventario(request, inventario_id):
    inventario = get_object_or_404(Inventario, id=inventario_id)
    if request.method == 'POST':
        inventario.delete()
        messages.success(request, 'Existencia de inventario eliminada exitosamente.')
        return redirect('inventario:listar_inventario')
    return render(request, 'inventario/inventario/eliminar.html', {'inventario': inventario})

# Vista para mostrar stock total por medicamento
@personal_medico_required
def stock_medicamentos(request):
    medicamentos_list = Medicamento.objects.select_related('categoria', 'proveedor').all().order_by('nombre')
    for medicamento in medicamentos_list:
        medicamento.stock = medicamento.stock_actual
        medicamento.estado = medicamento.estado_stock
    
    paginator = Paginator(medicamentos_list, 10)
    page_number = request.GET.get('page')
    medicamentos = paginator.get_page(page_number)
    
    return render(request, 'inventario/stock_medicamentos.html', {'medicamentos': medicamentos})

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

# --- Vistas de Exportación --- #

@personal_medico_required
def exportar_stock_pdf(request):
    medicamentos = Medicamento.objects.select_related('categoria', 'proveedor').all().order_by('nombre')
    logo_path = str(BASE_DIR / 'static/img/logo.png')
    
    context = {
        'medicamentos': medicamentos,
        'logo_path': logo_path,
        'generation_date': datetime.datetime.now().strftime('%d/%m/%Y %H:%M')
    }

    template = get_template('inventario/pdf/stock_template.html')
    html = template.render(context)
    result = BytesIO()
    pdf = pisa.pisaDocument(BytesIO(html.encode("UTF-8")), result)

    if not pdf.err:
        response = HttpResponse(result.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="reporte_stock_{}.pdf"'.format(datetime.datetime.now().strftime("%Y%m%d"))
        return response
    
    return HttpResponse("Error al generar el PDF.", status=400)

@personal_medico_required
def exportar_medicamentos_pdf(request):
    medicamentos = Medicamento.objects.select_related('categoria', 'proveedor').all().order_by('nombre')
    logo_path = str(BASE_DIR / 'static/img/logo.png')
    
    context = {
        'medicamentos': medicamentos,
        'logo_path': logo_path,
        'generation_date': datetime.datetime.now().strftime('%d/%m/%Y %H:%M')
    }

    template = get_template('inventario/pdf/medicamentos_template.html')
    html = template.render(context)
    result = BytesIO()
    pdf = pisa.pisaDocument(BytesIO(html.encode("UTF-8")), result)

    if not pdf.err:
        response = HttpResponse(result.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="reporte_medicamentos_{}.pdf"'.format(datetime.datetime.now().strftime("%Y%m%d"))
        return response
    
    return HttpResponse("Error al generar el PDF.", status=400)

@personal_medico_required
def exportar_proveedores_pdf(request):
    proveedores = Proveedor.objects.all().order_by('nombre')
    logo_path = str(BASE_DIR / 'static/img/logo.png')
    
    context = {
        'proveedores': proveedores,
        'logo_path': logo_path,
        'generation_date': datetime.datetime.now().strftime('%d/%m/%Y %H:%M')
    }

    template = get_template('inventario/pdf/proveedores_template.html')
    html = template.render(context)
    result = BytesIO()
    pdf = pisa.pisaDocument(BytesIO(html.encode("UTF-8")), result)

    if not pdf.err:
        response = HttpResponse(result.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="reporte_proveedores_{}.pdf"'.format(datetime.datetime.now().strftime("%Y%m%d"))
        return response
    
    return HttpResponse("Error al generar el PDF.", status=400)

@personal_medico_required
def exportar_categorias_pdf(request):
    categorias = Categoria.objects.all().order_by('nombre')
    logo_path = str(BASE_DIR / 'static/img/logo.png')
    
    context = {
        'categorias': categorias,
        'logo_path': logo_path,
        'generation_date': datetime.datetime.now().strftime('%d/%m/%Y %H:%M')
    }

    template = get_template('inventario/pdf/categorias_template.html')
    html = template.render(context)
    result = BytesIO()
    pdf = pisa.pisaDocument(BytesIO(html.encode("UTF-8")), result)

    if not pdf.err:
        response = HttpResponse(result.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="reporte_categorias_{}.pdf"'.format(datetime.datetime.now().strftime("%Y%m%d"))
        return response
    
    return HttpResponse("Error al generar el PDF.", status=400)

@personal_medico_required
def exportar_inventario_pdf(request):
    inventario = Inventario.objects.select_related('medicamento').all().order_by('-created_at')
    logo_path = str(BASE_DIR / 'static/img/logo.png')
    
    context = {
        'inventario': inventario,
        'logo_path': logo_path,
        'generation_date': datetime.datetime.now().strftime('%d/%m/%Y %H:%M')
    }

    template = get_template('inventario/pdf/inventario_template.html')
    html = template.render(context)
    result = BytesIO()
    pdf = pisa.pisaDocument(BytesIO(html.encode("UTF-8")), result)

    if not pdf.err:
        response = HttpResponse(result.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="reporte_inventario_{}.pdf"'.format(datetime.datetime.now().strftime("%Y%m%d"))
        return response
    
    return HttpResponse("Error al generar el PDF.", status=400)

@personal_medico_required
def exportar_stock_excel(request):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="reporte_stock_{}.xlsx"'.format(datetime.datetime.now().strftime("%Y%m%d"))
    wb = Workbook()
    ws = wb.active
    ws.title = "Stock de Medicamentos"
    ws.freeze_panes = 'A2'
    headers = ['Código', 'Medicamento', 'Categoría', 'Stock Actual', 'Stock Mínimo', 'Estado']
    ws.append(headers)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="198754", end_color="198754", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    medicamentos = Medicamento.objects.all().order_by('nombre')
    for med in medicamentos:
        ws.append([med.codigo, med.nombre, med.categoria.nombre, med.stock_actual, med.stock_minimo, med.estado_stock])
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except: pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width
    wb.save(response)
    return response

@personal_medico_required
def exportar_medicamentos_excel(request):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="listado_medicamentos_{}.xlsx"'.format(datetime.datetime.now().strftime("%Y%m%d"))
    wb = Workbook()
    ws = wb.active
    ws.title = "Medicamentos"
    ws.freeze_panes = 'A2'
    headers = ['Código', 'Nombre', 'Descripción', 'Categoría', 'Proveedor', 'Precio Unitario', 'Stock Mínimo']
    ws.append(headers)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="0d6efd", end_color="0d6efd", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    medicamentos = Medicamento.objects.select_related('categoria', 'proveedor').all().order_by('nombre')
    for med in medicamentos:
        ws.append([med.codigo, med.nombre, med.descripcion, med.categoria.nombre, med.proveedor.nombre, med.precio_unitario, med.stock_minimo])
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except: pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width
    wb.save(response)
    return response

@personal_medico_required
def exportar_proveedores_excel(request):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="listado_proveedores_{}.xlsx"'.format(datetime.datetime.now().strftime("%Y%m%d"))
    wb = Workbook()
    ws = wb.active
    ws.title = "Proveedores"
    ws.freeze_panes = 'A2'
    headers = ['Nombre', 'Contacto', 'Teléfono', 'Email', 'Dirección']
    ws.append(headers)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="0d6efd", end_color="0d6efd", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    proveedores = Proveedor.objects.all().order_by('nombre')
    for p in proveedores:
        ws.append([p.nombre, p.contacto, p.telefono, p.email, p.direccion])
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except: pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width
    wb.save(response)
    return response

@personal_medico_required
def exportar_categorias_excel(request):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="listado_categorias_{}.xlsx"'.format(datetime.datetime.now().strftime("%Y%m%d"))
    wb = Workbook()
    ws = wb.active
    ws.title = "Categorías"
    ws.freeze_panes = 'A2'
    headers = ['Nombre', 'Descripción']
    ws.append(headers)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="0d6efd", end_color="0d6efd", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    categorias = Categoria.objects.all().order_by('nombre')
    for cat in categorias:
        ws.append([cat.nombre, cat.descripcion])
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except: pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width
    wb.save(response)
    return response

@personal_medico_required
def exportar_inventario_excel(request):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="reporte_inventario_{}.xlsx"'.format(datetime.datetime.now().strftime("%Y%m%d"))
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventario"
    ws.freeze_panes = 'A2'
    headers = ['Medicamento', 'Código', 'Lote', 'Cantidad', 'Fecha de Ingreso', 'Fecha de Caducidad', 'Estado']
    ws.append(headers)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="0d6efd", end_color="0d6efd", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    inventario = Inventario.objects.select_related('medicamento').all().order_by('-created_at')
    for item in inventario:
        ws.append([
            item.medicamento.nombre,
            item.medicamento.codigo,
            item.lote or 'N/A',
            item.cantidad,
            item.created_at.strftime('%d/%m/%Y'),
            item.fecha_caducidad.strftime('%d/%m/%Y') if item.fecha_caducidad else 'N/A',
            item.estado
        ])
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except: pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width
    wb.save(response)
    return response


# Vistas para Movimientos
@personal_medico_required
def listar_movimientos(request):
    movimientos_list = MovimientoInventario.objects.select_related('medicamento').all().order_by('-fecha')
    paginator = Paginator(movimientos_list, 10)
    page_number = request.GET.get('page')
    movimientos = paginator.get_page(page_number)
    return render(request, 'inventario/movimientos/listar.html', {'movimientos': movimientos})

@personal_medico_required
def crear_salida_inventario(request):
    if request.method == 'POST':
        form = MovimientoSalidaForm(request.POST)
        if form.is_valid():
            medicamento = form.cleaned_data['medicamento']
            cantidad = form.cleaned_data['cantidad']
            descripcion = form.cleaned_data['descripcion']

            MovimientoInventario.objects.create(
                medicamento=medicamento,
                tipo='salida',
                cantidad=cantidad,
                descripcion=descripcion,
                usuario=request.user.username
            )
            
            messages.success(request, f'Salida de {cantidad} unidad(es) de {medicamento.nombre} registrada exitosamente.')
            return redirect('inventario:listar_movimientos')
    else:
        form = MovimientoSalidaForm()
    
    return render(request, 'inventario/movimientos/crear_salida.html', {
        'form': form,
        'titulo': 'Registrar Salida de Medicamento'
    })


# Vista principal del inventario
@personal_medico_required
def index(request):
    return render(request, 'inventario/index.html')

# Vista para creación de medicamentos vía AJAX
@login_required
@require_POST
def crear_medicamento_ajax(request):
    data = request.POST.copy()

    # Manejar creación de categoría al vuelo
    categoria_val = data.get('categoria')
    if categoria_val and not categoria_val.isdigit():
        categoria, _ = Categoria.objects.get_or_create(nombre=categoria_val.strip())
        data['categoria'] = categoria.id

    # Manejar creación de proveedor al vuelo
    proveedor_val = data.get('proveedor')
    if proveedor_val and not proveedor_val.isdigit():
        proveedor, _ = Proveedor.objects.get_or_create(nombre=proveedor_val.strip())
        data['proveedor'] = proveedor.id

    form = MedicamentoModalForm(data)
    if form.is_valid():
        medicamento = form.save()
        return JsonResponse({'success': True, 'id': medicamento.id, 'nombre': medicamento.nombre})
    else:
        errors = {field: error[0] for field, error in form.errors.items()}
        return JsonResponse({'success': False, 'errors': errors}, status=400)

@login_required
@require_POST
def crear_categoria_ajax(request):
    try:
        data = json.loads(request.body)
        form = CategoriaModalForm(data)
        if form.is_valid():
            categoria = form.save()
            return JsonResponse({'success': True, 'id': categoria.id, 'nombre': categoria.nombre})
        else:
            errors = {field: error[0] for field, error in form.errors.items()}
            return JsonResponse({'success': False, 'errors': errors}, status=400)
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'errors': 'Invalid JSON'}, status=400)

@login_required
@require_POST
def crear_proveedor_ajax(request):
    try:
        data = json.loads(request.body)
        form = ProveedorModalForm(data)
        if form.is_valid():
            proveedor = form.save()
            return JsonResponse({'success': True, 'id': proveedor.id, 'nombre': proveedor.nombre})
        else:
            errors = {field: error[0] for field, error in form.errors.items()}
            return JsonResponse({'success': False, 'errors': errors}, status=400)
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'errors': 'Invalid JSON'}, status=400)