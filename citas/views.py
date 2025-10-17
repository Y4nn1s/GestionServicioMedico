from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.core.paginator import Paginator
from django.utils import timezone
from django.db.models import Q, Count
from django.db import transaction
from django.core.exceptions import ValidationError
from django.db import IntegrityError
from django.http import JsonResponse
from django.views.decorators.http import require_http_methods
from django.views.decorators.csrf import csrf_exempt
import json
from django.http import HttpResponse
from django.template.loader import get_template
from xhtml2pdf import pisa
from io import BytesIO
import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from sistema_medico.settings import BASE_DIR

from .models import Cita, EstadoCita, TipoCita, MotivoCita, NotaCita, TipoNota
from .forms import CitaForm, EstadoCitaForm, TipoCitaForm, MotivoCitaForm
from pacientes.models import Paciente

def index(request):
    # Obtener los parámetros de la URL
    estado_filtro = request.GET.get('estado')
    query = request.GET.get('q')
    
    # Construir la consulta base para todas las citas
    citas_list = Cita.objects.all().select_related('paciente', 'tipo_cita', 'motivo', 'estado').order_by('-fecha', '-hora_inicio')
    
    # Aplicar filtro por estado si se especifica
    if estado_filtro:
        if estado_filtro == 'pendientes':
            citas_list = citas_list.filter(estado__nombre='Programada')
        elif estado_filtro == 'completadas':
            citas_list = citas_list.filter(estado__nombre='Completada')
        elif estado_filtro == 'canceladas':
            citas_list = citas_list.filter(estado__nombre='Cancelada')
    
    # Aplicar búsqueda si se especifica
    if query:
        citas_list = citas_list.filter(
            Q(paciente__nombre__icontains=query) |
            Q(paciente__apellido__icontains=query) |
            Q(tipo_cita__nombre__icontains=query) |
            Q(motivo__nombre__icontains=query)
        )
    
    # Paginación
    paginator = Paginator(citas_list, 10)  # Mostrar 10 citas por página
    page_number = request.GET.get('page')
    citas = paginator.get_page(page_number)
    
    # Obtener citas de hoy
    hoy = timezone.now().date()
    citas_hoy = Cita.objects.filter(fecha=hoy).select_related('paciente', 'tipo_cita', 'motivo', 'estado').order_by('hora_inicio')
    
    # Obtener estadísticas
    total_citas = Cita.objects.count()
    citas_pendientes = Cita.objects.filter(estado__nombre='Programada').count()
    citas_completadas = Cita.objects.filter(estado__nombre='Completada').count()
    citas_canceladas = Cita.objects.filter(estado__nombre='Cancelada').count()
    
    context = {
        'citas': citas,
        'citas_hoy': citas_hoy,
        'total_citas': total_citas,
        'citas_pendientes': citas_pendientes,
        'citas_completadas': citas_completadas,
        'citas_canceladas': citas_canceladas,
        'hoy': hoy,
        'query': query,
    }
    
    return render(request, 'citas/index.html', context)

@transaction.atomic
def create(request):
    if request.method == 'POST':
        form = CitaForm(request.POST)
        if form.is_valid():
            try:
                cita = form.save()
                
                # Crear nota automática de creación
                tipo_nota, created = TipoNota.objects.get_or_create(
                    nombre='Sistema',
                    defaults={'descripcion': 'Notas generadas automáticamente por el sistema'}
                )
                
                NotaCita.objects.create(
                    cita=cita,
                    tipo_nota=tipo_nota,
                    contenido=f"Cita creada el {timezone.now().strftime('%Y-%m-%d %H:%M')}"
                )
                
                messages.success(request, 'Cita creada correctamente.')
                return redirect('citas:index')
                
            except ValidationError as e:
                messages.error(request, f'Error de validación: {", ".join(e.messages)}')
            except IntegrityError as e:
                messages.error(request, 'Error de integridad de datos. La cita podría solaparse con otra existente.')
            except Exception as e:
                messages.error(request, f'Error inesperado al crear la cita: {str(e)}')
                # Re-lanzar la excepción para que se revierta la transacción
                raise
        else:
            messages.error(request, 'Por favor, corrija los errores en el formulario.')
    else:
        form = CitaForm()
    
    # Obtener datos para selects
    pacientes = Paciente.objects.all().order_by('apellido', 'nombre')
    estados = EstadoCita.objects.all()
    tipos = TipoCita.objects.all()
    motivos = MotivoCita.objects.all()
    
    return render(request, 'citas/create.html', {
        'form': form,
        'pacientes': pacientes,
        'estados': estados,
        'tipos': tipos,
        'motivos': motivos,
    })

# Vistas AJAX para crear tipos, motivos y estados desde el formulario
@require_http_methods(["POST"])
@csrf_exempt
def crear_tipo_cita_ajax(request):
    try:
        data = json.loads(request.body)
        form = TipoCitaForm(data)
        if form.is_valid():
            tipo = form.save()
            return JsonResponse({
                'success': True,
                'id': tipo.id,
                'nombre': tipo.nombre
            })
        else:
            return JsonResponse({
                'success': False,
                'errors': form.errors
            })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        })

@require_http_methods(["POST"])
@csrf_exempt
def crear_motivo_cita_ajax(request):
    try:
        data = json.loads(request.body)
        form = MotivoCitaForm(data)
        if form.is_valid():
            motivo = form.save()
            return JsonResponse({
                'success': True,
                'id': motivo.id,
                'nombre': motivo.nombre
            })
        else:
            return JsonResponse({
                'success': False,
                'errors': form.errors
            })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        })

@require_http_methods(["POST"])
@csrf_exempt
def crear_estado_cita_ajax(request):
    try:
        data = json.loads(request.body)
        form = EstadoCitaForm(data)
        if form.is_valid():
            estado = form.save()
            return JsonResponse({
                'success': True,
                'id': estado.id,
                'nombre': estado.nombre
            })
        else:
            return JsonResponse({
                'success': False,
                'errors': form.errors
            })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        })

def show(request, cita_id):
    try:
        cita = get_object_or_404(Cita, id=cita_id)
        notas = cita.notas.all().select_related('tipo_nota')
        return render(request, 'citas/show.html', {
            'cita': cita,
            'notas': notas
        })
    except Exception as e:
        messages.error(request, f'Error al cargar la cita: {str(e)}')
        return redirect('citas:index')

@transaction.atomic
def edit(request, cita_id):
    cita = get_object_or_404(Cita, id=cita_id)
    
    if request.method == 'POST':
        form = CitaForm(request.POST, instance=cita)
        if form.is_valid():
            try:
                form.save()
                
                # Crear nota de edición
                tipo_nota, created = TipoNota.objects.get_or_create(
                    nombre='Sistema',
                    defaults={'descripcion': 'Notas generadas automáticamente por el sistema'}
                )
                
                NotaCita.objects.create(
                    cita=cita,
                    tipo_nota=tipo_nota,
                    contenido=f"Cita modificada el {timezone.now().strftime('%Y-%m-%d %H:%M')}"
                )
                
                messages.success(request, 'Cita actualizada correctamente.')
                return redirect('citas:index')
                
            except ValidationError as e:
                messages.error(request, f'Error de validación: {", ".join(e.messages)}')
            except IntegrityError as e:
                messages.error(request, 'Error de integridad de datos. La cita podría solaparse con otra existente.')
            except Exception as e:
                messages.error(request, f'Error inesperado al actualizar la cita: {str(e)}')
                raise
        else:
            messages.error(request, 'Por favor, corrija los errores en el formulario.')
    else:
        form = CitaForm(instance=cita)
    
    # Obtener datos para selects
    pacientes = Paciente.objects.all().order_by('apellido', 'nombre')
    estados = EstadoCita.objects.all()
    tipos = TipoCita.objects.all()
    motivos = MotivoCita.objects.all()
    
    return render(request, 'citas/edit.html', {
        'form': form,
        'cita': cita,
        'pacientes': pacientes,
        'estados': estados,
        'tipos': tipos,
        'motivos': motivos,
    })

@transaction.atomic
def destroy(request, cita_id):
    cita = get_object_or_404(Cita, id=cita_id)
    
    if request.method == 'POST':
        try:
            # Guardar información para el mensaje antes de eliminar
            cita_info = f"Cita de {cita.paciente} el {cita.fecha}"
            cita.delete()
            
            messages.success(request, f'{cita_info} eliminada correctamente.')
            return redirect('citas:index')
            
        except Exception as e:
            messages.error(request, f'Error al eliminar la cita: {str(e)}')
            return redirect('citas:show', cita_id=cita_id)
    
    return render(request, 'citas/destroy.html', {'cita': cita})

def search(request):
    query = request.GET.get('q', '')
    citas = Cita.objects.all().select_related('paciente', 'tipo_cita', 'motivo', 'estado').order_by('-fecha', '-hora_inicio')
    
    if query:
        citas = citas.filter(
            Q(paciente__nombre__icontains=query) |
            Q(paciente__apellido__icontains=query) |
            Q(motivo__nombre__icontains=query)
        )
    
    paginator = Paginator(citas, 10)
    page_number = request.GET.get('page')
    citas_page = paginator.get_page(page_number)
    
    return render(request, 'citas/search.html', {
        'citas': citas_page,
        'query': query
    })

def citas_hoy(request):
    try:
        hoy = timezone.now().date()
        citas = Cita.objects.filter(fecha=hoy).select_related(
            'paciente', 'estado', 'motivo'
        ).order_by('hora_inicio')
        
        return render(request, 'citas/hoy.html', {
            'citas': citas,
            'hoy': hoy
        })
    except Exception as e:
        messages.error(request, f'Error al cargar las citas de hoy: {str(e)}')
        return redirect('citas:index')

@transaction.atomic
def cambiar_estado(request, cita_id, estado_id):
    try:
        cita = get_object_or_404(Cita, id=cita_id)
        estado = get_object_or_404(EstadoCita, id=estado_id)
        
        # Guardar el estado anterior para la nota
        estado_anterior = cita.estado.nombre
        
        cita.estado = estado
        cita.save()
        
        # Crear nota de cambio de estado
        tipo_nota, created = TipoNota.objects.get_or_create(
            nombre='Sistema',
            defaults={'descripcion': 'Notas generadas automáticamente por el sistema'}
        )
        
        NotaCita.objects.create(
            cita=cita,
            tipo_nota=tipo_nota,
            contenido=f"Estado cambiado de '{estado_anterior}' a '{estado.nombre}' el {timezone.now().strftime('%Y-%m-%d %H:%M')}"
        )
        
        messages.success(request, f'Estado de cita actualizado a {estado.nombre}.')
        
    except ValidationError as e:
        messages.error(request, f'Error de validación: {", ".join(e.messages)}')
    except IntegrityError as e:
        messages.error(request, 'Error de integridad de datos al cambiar el estado.')
    except Exception as e:
        messages.error(request, f'Error inesperado al cambiar el estado: {str(e)}')
    
    return redirect('citas:show', cita_id=cita_id)

# --- Vistas de Exportación --- #

def exportar_citas_pdf(request):
    citas = Cita.objects.all().select_related('paciente', 'tipo_cita', 'motivo', 'estado').order_by('-fecha', '-hora_inicio')
    logo_path = str(BASE_DIR / 'static/img/logo.png')
    
    context = {
        'citas': citas,
        'logo_path': logo_path,
        'generation_date': datetime.datetime.now().strftime('%d/%m/%Y %H:%M')
    }

    template = get_template('citas/pdf_template.html')
    html = template.render(context)
    result = BytesIO()
    pdf = pisa.pisaDocument(BytesIO(html.encode("UTF-8")), result)

    if not pdf.err:
        response = HttpResponse(result.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="reporte_citas_{}.pdf"'.format(datetime.datetime.now().strftime("%Y%m%d"))
        return response
    
    return HttpResponse("Error al generar el PDF.", status=400)

def exportar_citas_excel(request):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="listado_citas_{}.xlsx"'.format(datetime.datetime.now().strftime("%Y%m%d"))
    wb = Workbook()
    ws = wb.active
    ws.title = "Citas"
    ws.freeze_panes = 'A2'
    headers = ['Paciente', 'Tipo de Cita', 'Motivo', 'Fecha', 'Hora Inicio', 'Hora Fin', 'Estado']
    ws.append(headers)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="0d6efd", end_color="0d6efd", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    citas = Cita.objects.all().select_related('paciente', 'tipo_cita', 'motivo', 'estado').order_by('-fecha', '-hora_inicio')
    for cita in citas:
        ws.append([
            f"{cita.paciente.nombre} {cita.paciente.apellido}",
            cita.tipo_cita.nombre,
            cita.motivo.nombre,
            cita.fecha,
            cita.hora_inicio,
            cita.hora_fin,
            cita.estado.nombre
        ])
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width
    wb.save(response)
    return response

# Vistas AJAX para crear tipos, motivos y estados de cita
@require_http_methods(["POST"])
@csrf_exempt
def crear_tipo_cita_ajax(request):
    try:
        data = json.loads(request.body)
        form = TipoCitaForm(data)
        if form.is_valid():
            tipo = form.save()
            return JsonResponse({
                'success': True,
                'id': tipo.id,
                'nombre': tipo.nombre
            })
        else:
            return JsonResponse({
                'success': False,
                'errors': form.errors
            })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        })

@require_http_methods(["POST"])
@csrf_exempt
def crear_motivo_cita_ajax(request):
    try:
        data = json.loads(request.body)
        form = MotivoCitaForm(data)
        if form.is_valid():
            motivo = form.save()
            return JsonResponse({
                'success': True,
                'id': motivo.id,
                'nombre': motivo.nombre
            })
        else:
            return JsonResponse({
                'success': False,
                'errors': form.errors
            })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        })

@require_http_methods(["POST"])
@csrf_exempt
def crear_estado_cita_ajax(request):
    try:
        data = json.loads(request.body)
        form = EstadoCitaForm(data)
        if form.is_valid():
            estado = form.save()
            return JsonResponse({
                'success': True,
                'id': estado.id,
                'nombre': estado.nombre
            })
        else:
            return JsonResponse({
                'success': False,
                'errors': form.errors
            })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        })