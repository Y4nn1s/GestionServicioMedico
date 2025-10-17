from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.core.paginator import Paginator
from django.db import transaction
from django.db.models import Q
from django.http import HttpResponse
from django.template.loader import get_template
from xhtml2pdf import pisa
from io import BytesIO
import datetime
from sistema_medico.settings import BASE_DIR

from django.contrib.auth.decorators import login_required

from core.decorators import medico_required

from .models import HistorialMedico, Alergia, Enfermedad
from inventario.models import Medicamento
from .forms import HistorialMedicoForm
from django.http import JsonResponse
from django.views.decorators.http import require_POST
import json

@medico_required
def index(request):
    historiales_list = HistorialMedico.objects.select_related('paciente').prefetch_related(
        'alergias', 'enfermedades_preexistentes', 'medicamentos_actuales'
    ).order_by('-updated_at')
    
    paginator = Paginator(historiales_list, 10)
    page_number = request.GET.get('page')
    historiales = paginator.get_page(page_number)
    
    return render(request, 'historiales/index.html', {'historiales': historiales})

from inventario.forms import MedicamentoModalForm

# ... (existing code) ...

@medico_required
@transaction.atomic
def create(request):
    if request.method == 'POST':
        form = HistorialMedicoForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Historial médico creado correctamente.')
            return redirect('historiales:index')
        else:
            messages.error(request, 'Por favor, corrija los errores en el formulario.')
    else:
        form = HistorialMedicoForm()
    
    medicamento_form = MedicamentoModalForm()
    return render(request, 'historiales/create.html', {'form': form, 'medicamento_form': medicamento_form})

@medico_required
def show(request, historial_id):
    historial = get_object_or_404(
        HistorialMedico.objects.select_related('paciente').prefetch_related(
            'alergias', 'enfermedades_preexistentes', 'medicamentos_actuales'
        ),
        id=historial_id
    )
    return render(request, 'historiales/show.html', {'historial': historial})

@medico_required
@transaction.atomic
def edit(request, historial_id):
    historial = get_object_or_404(HistorialMedico, id=historial_id)
    
    if request.method == 'POST':
        form = HistorialMedicoForm(request.POST, instance=historial)
        if form.is_valid():
            form.save()
            messages.success(request, 'Historial médico actualizado correctamente.')
            return redirect('historiales:index')
        else:
            messages.error(request, 'Por favor, corrija los errores en el formulario.')
    else:
        form = HistorialMedicoForm(instance=historial)
    
    medicamento_form = MedicamentoModalForm()
    return render(request, 'historiales/edit.html', {'form': form, 'historial': historial, 'medicamento_form': medicamento_form})

@medico_required
@transaction.atomic
def destroy(request, historial_id):
    historial = get_object_or_404(HistorialMedico, id=historial_id)
    
    if request.method == 'POST':
        paciente_info = str(historial.paciente)
        historial.delete()
        messages.success(request, f'Historial médico de {paciente_info} eliminado correctamente.')
        return redirect('historiales:index')
    
    return render(request, 'historiales/destroy.html', {'historial': historial})

@medico_required
def search(request):
    query = request.GET.get('q', '')
    historiales = HistorialMedico.objects.select_related('paciente').order_by('-updated_at')
    
    if query:
        historiales = historiales.filter(
            Q(paciente__nombre__icontains=query) |
            Q(paciente__apellido__icontains=query) |
            Q(paciente__numero_documento__icontains=query) |
            Q(alergias__nombre__icontains=query) |
            Q(enfermedades_preexistentes__nombre__icontains=query)
        ).distinct()
    
    paginator = Paginator(historiales, 10)
    page_number = request.GET.get('page')
    historiales_page = paginator.get_page(page_number)
    
    return render(request, 'historiales/search.html', {
        'historiales': historiales_page,
        'query': query
    })

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

import os

# --- Vistas de Exportación --- #

@medico_required
def exportar_historial_individual_pdf(request, historial_id):
    historial = get_object_or_404(HistorialMedico, id=historial_id)
    logo_path = str(BASE_DIR / 'static/img/logo.png')
    
    context = {
        'historial': historial,
        'logo_path': logo_path,
        'generation_date': datetime.datetime.now().strftime('%d/%m/%Y %H:%M')
    }

    template = get_template('historiales/pdf_template.html')
    html = template.render(context)
    result = BytesIO()
    pdf = pisa.pisaDocument(BytesIO(html.encode("UTF-8")), result)

    if not pdf.err:
        response = HttpResponse(result.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="historial_{}_{}.pdf"'.format(historial.paciente.numero_documento, datetime.datetime.now().strftime("%Y%m%d"))
        return response
    
    return HttpResponse("Error al generar el PDF.", status=400)

@medico_required
def exportar_historiales_pdf(request):
    historiales = HistorialMedico.objects.select_related('paciente').prefetch_related('alergias', 'enfermedades_preexistentes', 'medicamentos_actuales').all()
    logo_path = str(BASE_DIR / 'static/img/logo.png')
    
    context = {
        'historiales': historiales,
        'logo_path': logo_path,
        'generation_date': datetime.datetime.now().strftime('%d/%m/%Y %H:%M')
    }

    template = get_template('historiales/pdf/historiales_template.html')
    html = template.render(context)
    result = BytesIO()
    pdf = pisa.pisaDocument(BytesIO(html.encode("UTF-8")), result)

    if not pdf.err:
        response = HttpResponse(result.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="reporte_historiales_{}.pdf"'.format(datetime.datetime.now().strftime("%Y%m%d"))
        return response
    
    return HttpResponse("Error al generar el PDF.", status=400)

@medico_required
def exportar_historiales_excel(request):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="listado_historiales_{}.xlsx"'.format(datetime.datetime.now().strftime("%Y%m%d"))
    wb = Workbook()
    ws = wb.active
    ws.title = "Historiales Médicos"
    ws.freeze_panes = 'A2'
    
    headers = ['Paciente', 'Cédula', 'Fecha de Creación', 'Última Actualización', 'Alergias', 'Enfermedades Preexistentes', 'Medicamentos Actuales']
    ws.append(headers)
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="0d6efd", end_color="0d6efd", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    historiales = HistorialMedico.objects.select_related('paciente').prefetch_related('alergias', 'enfermedades_preexistentes', 'medicamentos_actuales').all()

    for historial in historiales:
        alergias = ", ".join([a.nombre for a in historial.alergias.all()]) or "N/A"
        enfermedades = ", ".join([e.nombre for e in historial.enfermedades_preexistentes.all()]) or "N/A"
        medicamentos = ", ".join([m.nombre for m in historial.medicamentos_actuales.all()]) or "N/A"
        
        ws.append([
            str(historial.paciente),
            historial.paciente.numero_documento,
            historial.created_at.strftime('%d/%m/%Y %H:%M'),
            historial.updated_at.strftime('%d/%m/%Y %H:%M'),
            alergias,
            enfermedades,
            medicamentos
        ])

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    wb.save(response)
    return response


# Vistas para AJAX
@login_required
@require_POST
def crear_alergia_ajax(request):
    try:
        data = json.loads(request.body)
        nombre = data.get('nombre')
        if not nombre:
            return JsonResponse({'success': False, 'errors': 'El nombre es requerido.'}, status=400)
        
        alergia, created = Alergia.objects.get_or_create(nombre=nombre.strip())
        
        if created:
            return JsonResponse({'success': True, 'id': alergia.id, 'nombre': alergia.nombre})
        else:
            return JsonResponse({'success': False, 'errors': 'La alergia ya existe.'}, status=400)
            
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'errors': 'JSON inválido.'}, status=400)

@login_required
@require_POST
def crear_enfermedad_ajax(request):
    try:
        data = json.loads(request.body)
        nombre = data.get('nombre')
        if not nombre:
            return JsonResponse({'success': False, 'errors': 'El nombre es requerido.'}, status=400)
        
        enfermedad, created = Enfermedad.objects.get_or_create(nombre=nombre.strip())
        
        if created:
            return JsonResponse({'success': True, 'id': enfermedad.id, 'nombre': enfermedad.nombre})
        else:
            return JsonResponse({'success': False, 'errors': 'La enfermedad ya existe.'}, status=400)

    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'errors': 'JSON inválido.'}, status=400)
